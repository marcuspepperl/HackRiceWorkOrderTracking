from openpyxl import load_workbook, Workbook
import math

def parse_date(date_str):

    date_time = date_str.split()
    if date_time[0].find(':') != -1:
        temp = date_time[0]
        date_time[0] = date_time[1]
        date_time[1] = temp

    date, time = date_time

    if date.find('/') != -1:
        month, day, year = date.split('/')
    elif date.find('-') != -1:
        year, month, day = date.split('-')

    if len(day) == 1:
        day = '0' + day

    if len(month) == 1:
        month = '0' + month

    if len(year) == 2:
        year = '20' + year

    return day, month, year

def distance(location1, location2):
    return Math.sqrt((location1[0] - location2[0]) ** 2 + (location1[1] - location2[1]) ** 2)


class Time:

    def __init__(self, hour, minutes):

        self.hour = hour
        self.minutes = minutes

    def add(self, other):
        other_hour = other.get_hour()
        other_minutes = other.get_minutes()
        combined_minutes = self.minutes + other_minutes
        added_hours = combined_minutes // 60 + other_hour
        added_minutes = combined_minutes % 60

        return Time(self.hour + added_hours, self.minutes + added_minutes)

    def subtract(self, other):
        if not self.greater(other):
            return -1

        other_hour = other.get_hour()
        other_minutes = other.get_minutes()
        difference_minutes = self.minutes - other_minutes
        resultant_minutes = difference_minutes
        resultant_hours = self.hour - other_hour

        if difference_minutes < 0:
            resultant_minutes += 60
            resultant_hours -= 1

        return Time(resultant_hours, resultant_minutes)


    def greater(self, other):
        return self.get_time() > other.get_time()

    def get_time(self):
        return self.hour, self.minutes

    def get_hour(self):
        return self.hour

    def get_minutes(self):
        return self.minutes


class Schedule:

    def __init__(self, begin_shift, end_shift):

        self.begin_shift = begin_shift
        self.end_shift = end_shift
        self.shift_time = self.end_shift.subtract(self.begin_shift)
        self.schedule = list()
        self.schedule.append(list())
        self.schedule_days = 0

    def add_task(self, order, time, current_location, order_location):

        if not self.available:
            return False

        if self.schedule_days == 1:
            last_end_time = self.schedule[0][-1][2] + movement_time

        else:
            last_end_time = self.begin_shift

        travel_distance = int(80 * distance(current_location, order_location))

        movement_time = Time(0, travel_distance * 2)

        start_time = last_end_time.add(movement_time)

        remaining_time = self.end_shift.subtract(start_time)

        if not time.greater(remaining_time):
            self.schedule[0].append((order, start_time, start_time.add(time)))

        else:
            overflow_time = time.subtract(remaining_time)
            self.schedule[0].append((order, start_time, self.end_shift))
            day = 1
            while overflow_time > 0:

                if overflow_time.greater(self.shift_time):
                    self.schedule[day].append((order, self.begin_shift, self.end_shift))
                    overflow_time = overflow_time.subtract(self.shift_time)

                else:
                    self.schedule[day].append((order, self.begin_shift, self.begin_shift.add(overflow_time)))
                    overflow_time = 0

        return True


    def is_available(self):
        return self.schedule_days < 2

    def get_time_remaining(self):
        if not self.schedule[0]:
            return self.shift_time
        else:
            return self.end_shift.subtract(self.schedule[0][-1][2])

    def update(self):
        self.schedule.pop(0)
        if not self.schedule:
            self.schedule.append(list())

class Facility_Schedule:

    def __init__(self, max_capacity):

        self.schedule = list()
        self.schedule.append(list())
        self.max_capacity = max_capacity

    def add_task(self, order, begin_time, end_time):







SHIFTS = {"Morning" : (Time(4, 0), Time(14, 0)), "Evening" : (Time(14, 0), Time(23, 0))}

class Read_Data():

    def __init__(self):

        self.NULL = {None, "=""", ""}
        self.WORKFILE = "data_file.xlsx"
        self.SHEET_NAMES = ["Equipment Details", "Worker Details", "Facility Details", "Work Order Examples"]
        self.wb = load_workbook(self.WORKFILE)

        self.EQUIPMENT_FIELDS = ["failure", "fix", "fac1", "fac2", "fac3", "fac4", "fac5"]
        self.WORKER_FIELDS = ["certifications", "shifts", "latitude", "longitude"]
        self.FACILITY_FIELDS = ["latitude", "longitude", "occupancy"]
        self.ORDER_FIELDS = ["facility", "type", "id", "priority", "completion_time", "submission_time"]

        self.EQUIPMENT = self.read_sheet(self.SHEET_NAMES[0], 3, 2, self.EQUIPMENT_FIELDS)
        self.WORKER = self.read_sheet(self.SHEET_NAMES[1], 2, 2, self.WORKER_FIELDS)
        self.FACILITY = self.read_sheet(self.SHEET_NAMES[2], 3, 2, self.FACILITY_FIELDS)
        self.ORDER = self.read_sheet(self.SHEET_NAMES[3], 3, 2, self.ORDER_FIELDS)

        self.DISTANCE_MATRIX = self.generate_distance_matrix()


    def read_sheet(self, sheetname, start_row, start_col, fields):

        data = dict()
        ws = self.wb[sheetname]
        row = start_row

        while True:
            key = ws.cell(row, start_col).value
            if key in self.NULL:
                break
            field_dict = dict()
            col = start_col + 1
            for index, field in enumerate(fields):
                field_dict[field] = str(ws.cell(row, col + index).value)
            data[key] = field_dict

            row += 1

        return data

    def generate_distance_matrix(self):

        n = len(self.FACILITY)
        coordinates = list()
        for fields in self.FACILITY.values():
            pair = (float(fields["latitude"]), float(fields["longitude"]))
            coordinates.append(pair)

        matrix = list()
        for index1 in range(n):
            distance = list()
            latitude1 = coordinates[index1][0]
            longitude1 = coordinates[index1][1]
            for index2 in range(n):
                latitude2 = coordinates[index2][0]
                longitude2 = coordinates[index2][1]
                distance.append(math.sqrt((latitude2 - latitude1) ** 2 + (longitude2 - longitude1) ** 2))

            matrix.append(distance)

        return matrix


    def get_equipment(self):
        return self.EQUIPMENT

    def get_worker(self):
        return self.WORKER

    def get_facility(self):
        return self.FACILITY

    def get_order(self):
        return self.ORDER


class Equipment():

    def __init__(self, key, equipment_fields):

        self.key = key
        self.failure = equipment_fields["failure"]
        self.fix = equipment_fields["fix"]
        self.fac1 = equipment_fields["fac1"]
        self.fac2 = equipment_fields["fac2"]
        self.fac3 = equipment_fields["fac3"]
        self.fac4 = equipment_fields["fac4"]
        self.fac5 = equipment_fields["fac5"]

    def get_key(self):
        return self.key

    def get_failure(self):
        return self.failure

    def get_fix(self):
        return self.fix

    def get_fac1(self):
        return self.fac1

    def get_fac2(self):
        return self.fac2

    def get_fac3(self):
        return self.fac3

    def get_fac4(self):
        return self.fac4

    def get_fac5(self):
        return self.fac5


class Worker():

    def __init__(self, key, worker_fields):

        self.key = key
        certifications = worker_fields["certifications"]
        self.certifications = set(certifications.split(", "))
        self.shifts = worker_fields["shifts"]
        self.latitude = worker_fields["latitude"]
        self.longitude = worker_fields["longitude"]

        self.start_shift, self.end_shift = SHIFTS[self.shifts]
        self.schedule = Schedule(self.start_shift, self.end_shift)

    def get_key(self):
        return self.key

    def get_certifications(self):
        return self.certifications

    def get_shifts(self):
        return self.shifts

    def get_latitude(self):
        return self.latitude

    def get_longitude(self):
        return self.longitude

    def get_schedule(self):
        return self.schedule


class Facility():

    def __init__(self, key, facility_fields):

        self.key = key
        self.latitude = facility_fields["latitude"]
        self.longitude = facility_fields["longitude"]
        self.occupancy = facility_fields["occupancy"]

    def get_key(self):
        return self.key

    def get_latitude(self):
        return self.latitude

    def get_longitude(self):
        return self.longitude

    def get_occupancy(self):
        return self.occupancy


class Order():

    def __init__(self, key, order_fields):

        self.key = key
        self.facility = order_fields["facility"]
        self.type = order_fields["type"]
        self.id = order_fields["id"]
        self.priority = order_fields["priority"]
        self.completion_time = order_fields["completion_time"]
        self.submission_time = order_fields["submission_time"]

    def get_key(self):
        return self.key

    def get_facility(self):
        return self.facility

    def get_type(self):
        return self.type

    def get_id(self):
        return self.id

    def get_priority(self):
        return self.priority

    def get_completion_time(self):
        return self.completion_time

    def get_submission_time(self):
        return self.submission_time



class Data():

    def __init__(self):

        self.data = Read_Data()
        self.equipment_data = self.data.get_equipment()
        self.worker_data = self.data.get_worker()
        self.facility_data = self.data.get_facility()
        self.order_data = self.data.get_order()

        self.equipment_objs = dict()
        self.worker_objs = dict()
        self.facility_objs = dict()
        self.order_objs = dict()

        for key, equipment_fields in self.equipment_data.items():
            self.equipment_objs[key] = Equipment(key, equipment_fields)

        for key, worker_fields in self.worker_data.items():
            self.worker_objs[key] = Worker(key, worker_fields)

        for key, facility_fields in self.facility_data.items():
            self.facility_objs[key] = Facility(key, facility_fields)

        for key, order_fields in self.order_data.items():
            self.order_objs[key] = Order(key, order_fields)

        distinct_days = set()
        order_days = list()

        for key, order in self.order_objs.items():

            submission_time = order.get_submission_time()
            order_day = int(parse_date(submission_time)[0])
            distinct_days.add(order_day)
            order_days.append((key, order_day))

        min_day = min(distinct_days)
        max_day = max(distinct_days)
        total_days = max_day - min_day + 1

        self.daily_order_objs = list()
        for _ in range(total_days):
            self.daily_order_objs.append(dict())
        for order in order_days:
            key, order_day = order
            day_offset = order_day - min_day
            self.daily_order_objs[day_offset][key] = self.get_order(key)


    def get_equipment(self, name):
        if name in self.equipment_objs:
            return self.equipment_objs[name]
        return None

    def get_worker(self, name):
        if name in self.worker_objs:
            return self.worker_objs[name]
        return None

    def get_facility(self, name):
        if name in self.facility_objs:
            return self.facility_objs[name]
        return None

    def get_order(self, name):
        if name in self.order_objs:
            return self.order_objs[name]
        return None

    def get_all_equipment(self):
        return self.equipment_objs

    def get_all_workers(self):
        return self.worker_objs

    def get_all_facilities(self):
        return self.facility_objs

    def get_all_orders(self):
        return self.order_objs

    def get_all_daily_orders(self):
        return self.daily_order_objs


def main():
    data = Data()
    print(data.get_worker("Bob").get_certifications())

if __name__ == '__main__':
    main()
